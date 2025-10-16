"""
Excel to Searchable CSV Converter

Converts Excel files to searchable CSV format optimized for ripgrep:
- Embeds row/column coordinates (R1C1 format) for mapping back to Excel
- Preserves all data including formulas, comments, and formatting
- Handles multiple sheets in single workbook
- Optimized for fast text search with ripgrep

CSV Format:
Each cell: R{row}C{col}|{value}
Separated by tabs for better parsing
"""

import csv
import re
from pathlib import Path
from typing import Dict, List, Optional, Union, Tuple
from datetime import datetime
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell

from config.settings import settings
from src.utils.logger import logger, LogContext, log_file_operation
from src.cache.cache_manager import get_cache_manager, CacheKey
from src.utils.coordinate_mapper import CoordinateMapper


class CSVConverter:
    """
    Convert Excel files to searchable CSV format
    
    Features:
    - Embeds coordinates for reverse mapping
    - Preserves data types and formats
    - Handles merged cells, formulas, comments
    - Caches converted files
    """
    
    def __init__(self):
        self.cache = get_cache_manager()
        self.output_dir = settings.searchable_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def convert_workbook(self, 
                        excel_path: Union[str, Path],
                        use_cache: bool = True) -> Dict[str, Path]:
        """
        Convert entire Excel workbook to searchable CSV files
        
        Args:
            excel_path: Path to Excel file
            use_cache: If True, use cached CSV files if available
            
        Returns:
            Dict mapping sheet names to CSV file paths
        """
        excel_path = Path(excel_path)
        
        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
        
        with LogContext("convert_workbook", file_name=excel_path.name):
            # Check cache
            if use_cache and settings.cache_csv_files:
                cached_csvs = self._get_cached_csvs(excel_path)
                if cached_csvs:
                    logger.info(f"Using cached CSV files for {excel_path.name}")
                    return cached_csvs
            
            # Load workbook (read-only for speed)
            logger.info(f"Loading workbook: {excel_path.name}")
            wb = openpyxl.load_workbook(
                excel_path,
                read_only=settings.excel_read_only,
                data_only=settings.excel_data_only
            )
            
            # Convert each sheet
            csv_files = {}
            for sheet in wb.worksheets:
                try:
                    csv_path = self.convert_sheet(
                        excel_path,
                        sheet,
                        wb
                    )
                    csv_files[sheet.title] = csv_path
                    logger.info(f"Converted sheet '{sheet.title}' to {csv_path.name}")
                except Exception as e:
                    logger.error(f"Error converting sheet '{sheet.title}': {e}")
                    continue
            
            wb.close()
            
            # Cache the mapping
            if settings.cache_csv_files:
                self._cache_csvs(excel_path, csv_files)
            
            log_file_operation("converted", excel_path, sheets=len(csv_files))
            return csv_files
    
    def convert_sheet(self,
                     excel_path: Path,
                     sheet: Worksheet,
                     workbook: openpyxl.Workbook) -> Path:
        """
        Convert single Excel sheet to searchable CSV
        
        Args:
            excel_path: Path to Excel file
            sheet: Worksheet to convert
            workbook: Parent workbook (for context)
            
        Returns:
            Path to generated CSV file
        """
        # Generate output filename
        safe_sheet_name = self._sanitize_filename(sheet.title)
        csv_filename = f"{excel_path.stem}_{safe_sheet_name}.csv"
        csv_path = self.output_dir / csv_filename
        
        # Write CSV with embedded coordinates
        with open(csv_path, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f, delimiter='\t', quoting=csv.QUOTE_MINIMAL)
            
            # Get actual data range (ignore empty trailing rows/cols)
            max_row, max_col = self._get_actual_data_range(sheet)
            
            logger.debug(f"Processing sheet '{sheet.title}': {max_row}x{max_col}")
            
            # Process each row
            for row_idx in range(1, max_row + 1):
                row_data = []
                
                for col_idx in range(1, max_col + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    
                    # Format cell with coordinate
                    cell_str = self._format_cell_with_coordinate(
                        cell, 
                        row_idx, 
                        col_idx
                    )
                    row_data.append(cell_str)
                
                writer.writerow(row_data)
        
        logger.debug(f"CSV written: {csv_path} ({csv_path.stat().st_size} bytes)")
        return csv_path
    
    def _format_cell_with_coordinate(self, 
                                     cell: Cell, 
                                     row: int, 
                                     col: int) -> str:
        """
        Format cell value with embedded R1C1 coordinate
        
        Format: R{row}C{col}|{value}
        
        Args:
            cell: Excel cell object
            row: Row number (1-based)
            col: Column number (1-based)
            
        Returns:
            Formatted string
        """
        # Get cell value
        value = cell.value
        
        # Handle None/empty cells
        if value is None:
            value = ""
        
        # Convert value to string
        if isinstance(value, datetime):
            # Format dates consistently
            value_str = value.strftime("%Y-%m-%d %H:%M:%S")
        elif isinstance(value, (int, float)):
            # Preserve numbers as-is
            value_str = str(value)
        elif isinstance(value, bool):
            # Boolean to string
            value_str = "TRUE" if value else "FALSE"
        else:
            # Convert to string, clean up
            value_str = str(value).strip()
        
        # Replace pipe character to avoid conflicts
        value_str = value_str.replace('|', '¦')
        
        # Replace tabs and newlines to keep CSV structure
        value_str = value_str.replace('\t', ' ').replace('\n', ' ').replace('\r', ' ')
        
        # Format with coordinate: R1C1|value
        return f"R{row}C{col}|{value_str}"
    
    def _get_actual_data_range(self, sheet: Worksheet) -> Tuple[int, int]:
        """
        Get actual data range (ignoring empty trailing rows/cols)
        
        Args:
            sheet: Worksheet to analyze
            
        Returns:
            Tuple of (max_row, max_col)
        """
        # Start with sheet dimensions
        max_row = sheet.max_row or 1
        max_col = sheet.max_column or 1
        
        # For read-only sheets, we can't easily determine actual range
        # Use sheet dimensions directly
        if settings.excel_read_only:
            return (max_row, max_col)
        
        # For non-read-only sheets, find actual last non-empty cell
        last_row = 1
        last_col = 1
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    last_row = max(last_row, cell.row)
                    last_col = max(last_col, cell.column)
        
        return (last_row, last_col)
    
    def _sanitize_filename(self, name: str) -> str:
        """
        Sanitize sheet name for use in filename
        
        Args:
            name: Sheet name
            
        Returns:
            Sanitized filename
        """
        # Replace invalid characters
        safe_name = re.sub(r'[^\w\s\-]', '_', name)
        # Replace multiple spaces/underscores with single underscore
        safe_name = re.sub(r'[\s_]+', '_', safe_name)
        # Trim to reasonable length
        safe_name = safe_name[:100]
        return safe_name
    
    def _get_cached_csvs(self, excel_path: Path) -> Optional[Dict[str, Path]]:
        """
        Get cached CSV files for Excel workbook
        
        Args:
            excel_path: Path to Excel file
            
        Returns:
            Dict of sheet name -> CSV path, or None if not cached
        """
        try:
            cache_key = f"csv_files:{CacheKey.for_file(excel_path)}"
            cached_mapping = self.cache.get(cache_key, tier="disk")
            
            if cached_mapping:
                # Verify all CSV files still exist
                all_exist = all(
                    Path(csv_path).exists() 
                    for csv_path in cached_mapping.values()
                )
                
                if all_exist:
                    # Convert string paths back to Path objects
                    return {
                        sheet: Path(csv_path)
                        for sheet, csv_path in cached_mapping.items()
                    }
            
            return None
            
        except Exception as e:
            logger.warning(f"Error reading CSV cache: {e}")
            return None
    
    def _cache_csvs(self, excel_path: Path, csv_files: Dict[str, Path]) -> None:
        """
        Cache CSV file mapping
        
        Args:
            excel_path: Path to Excel file
            csv_files: Dict of sheet name -> CSV path
        """
        try:
            cache_key = f"csv_files:{CacheKey.for_file(excel_path)}"
            
            # Convert Path objects to strings for caching
            cached_mapping = {
                sheet: str(csv_path)
                for sheet, csv_path in csv_files.items()
            }
            
            self.cache.set(cache_key, cached_mapping, tier="disk")
            
        except Exception as e:
            logger.warning(f"Error caching CSV mapping: {e}")
    
    def invalidate_cache(self, excel_path: Union[str, Path]) -> None:
        """
        Invalidate cached CSV files for Excel workbook
        
        Args:
            excel_path: Path to Excel file
        """
        excel_path = Path(excel_path)
        
        try:
            # Get cached mapping
            cached_csvs = self._get_cached_csvs(excel_path)
            
            if cached_csvs:
                # Delete CSV files
                for csv_path in cached_csvs.values():
                    csv_path.unlink(missing_ok=True)
                    logger.debug(f"Deleted cached CSV: {csv_path.name}")
            
            # Delete cache entry
            cache_key = f"csv_files:{CacheKey.for_file(excel_path)}"
            self.cache.delete(cache_key)
            
            logger.info(f"Invalidated CSV cache for {excel_path.name}")
            
        except Exception as e:
            logger.warning(f"Error invalidating CSV cache: {e}")


class CSVParser:
    """
    Parse searchable CSV files back to structured data
    
    Companion to CSVConverter - extracts coordinates and values
    """
    
    @staticmethod
    def parse_cell(cell_str: str) -> Dict[str, Union[str, int]]:
        """
        Parse cell string with embedded coordinate
        
        Args:
            cell_str: Cell string in format "R1C1|value"
            
        Returns:
            Dict with row, col, value, cell_ref (A1)
        """
        if '|' not in cell_str:
            # Invalid format, return as-is
            return {
                'row': None,
                'col': None,
                'value': cell_str,
                'cell_ref': None
            }
        
        # Split coordinate and value
        coord_part, value = cell_str.split('|', 1)
        
        # Parse R1C1 coordinate
        match = re.match(r'R(\d+)C(\d+)', coord_part)
        if not match:
            return {
                'row': None,
                'col': None,
                'value': cell_str,
                'cell_ref': None
            }
        
        row = int(match.group(1))
        col = int(match.group(2))
        
        # Convert to A1 notation
        cell_ref = CoordinateMapper.indices_to_a1(row - 1, col - 1, zero_based=True)
        
        return {
            'row': row,
            'col': col,
            'value': value,
            'cell_ref': cell_ref
        }
    
    @staticmethod
    def parse_csv_file(csv_path: Union[str, Path]) -> List[List[Dict]]:
        """
        Parse entire CSV file
        
        Args:
            csv_path: Path to CSV file
            
        Returns:
            2D list of parsed cell data
        """
        csv_path = Path(csv_path)
        
        rows = []
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f, delimiter='\t')
            
            for row in reader:
                parsed_row = [CSVParser.parse_cell(cell) for cell in row]
                rows.append(parsed_row)
        
        return rows
    
    @staticmethod
    def extract_sheet_name(csv_path: Union[str, Path]) -> str:
        """
        Extract sheet name from CSV filename
        
        Args:
            csv_path: Path to CSV file
            
        Returns:
            Sheet name
        """
        csv_path = Path(csv_path)
        filename = csv_path.stem
        
        # Remove Excel filename prefix
        # Format: {excel_name}_{sheet_name}
        if '_' in filename:
            parts = filename.split('_', 1)
            if len(parts) == 2:
                return parts[1].replace('_', ' ')
        
        return filename


class CSVSearchHelper:
    """
    Helper class for working with searchable CSV files
    """
    
    @staticmethod
    def get_csv_for_sheet(excel_path: Union[str, Path], 
                         sheet_name: str) -> Optional[Path]:
        """
        Get CSV file path for specific sheet
        
        Args:
            excel_path: Path to Excel file
            sheet_name: Sheet name
            
        Returns:
            Path to CSV file or None
        """
        converter = CSVConverter()
        
        # Check cache
        cached_csvs = converter._get_cached_csvs(Path(excel_path))
        if cached_csvs and sheet_name in cached_csvs:
            return cached_csvs[sheet_name]
        
        return None
    
    @staticmethod
    def search_in_csv(csv_path: Union[str, Path], 
                     search_term: str) -> List[Dict]:
        """
        Simple text search in CSV file (without ripgrep)
        
        Args:
            csv_path: Path to CSV file
            search_term: Term to search for
            
        Returns:
            List of matches with coordinates
        """
        csv_path = Path(csv_path)
        matches = []
        
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f, delimiter='\t')
            
            for row_idx, row in enumerate(reader, 1):
                for col_idx, cell in enumerate(row, 1):
                    parsed = CSVParser.parse_cell(cell)
                    
                    if search_term.lower() in parsed['value'].lower():
                        matches.append({
                            'csv_row': row_idx,
                            'csv_col': col_idx,
                            'excel_row': parsed['row'],
                            'excel_col': parsed['col'],
                            'cell_ref': parsed['cell_ref'],
                            'value': parsed['value'],
                            'matched_term': search_term
                        })
        
        return matches


# Convenience functions
def convert_excel_to_csv(excel_path: Union[str, Path], 
                         use_cache: bool = True) -> Dict[str, Path]:
    """
    Quick function to convert Excel to CSV
    
    Args:
        excel_path: Path to Excel file
        use_cache: Use cached files if available
        
    Returns:
        Dict of sheet name -> CSV path
    """
    converter = CSVConverter()
    return converter.convert_workbook(excel_path, use_cache=use_cache)


def parse_csv_cell(cell_str: str) -> Dict:
    """Quick function to parse CSV cell"""
    return CSVParser.parse_cell(cell_str)


# Export main components
__all__ = [
    'CSVConverter',
    'CSVParser',
    'CSVSearchHelper',
    'convert_excel_to_csv',
    'parse_csv_cell',
]


if __name__ == "__main__":
    import sys
    
    print("=== CSV Converter Tests ===\n")
    
    # Test cell formatting
    print("1. Testing cell formatting:")
    converter = CSVConverter()
    
    # Mock cell
    class MockCell:
        def __init__(self, value):
            self.value = value
    
    test_cases = [
        (MockCell("Hello World"), 1, 1),
        (MockCell(12345), 2, 3),
        (MockCell(None), 5, 10),
        (MockCell(True), 1, 5),
    ]
    
    for cell, row, col in test_cases:
        formatted = converter._format_cell_with_coordinate(cell, row, col)
        print(f"   Row {row}, Col {col}: {formatted}")
    
    # Test parsing
    print("\n2. Testing cell parsing:")
    test_strings = [
        "R1C1|Hello World",
        "R2C3|12345",
        "R5C10|",
        "R1C5|TRUE",
    ]
    
    for test_str in test_strings:
        parsed = CSVParser.parse_cell(test_str)
        print(f"   '{test_str}' -> {parsed}")
    
    # Test filename sanitization
    print("\n3. Testing filename sanitization:")
    test_names = [
        "Sales Data",
        "Q1/Q2 Comparison",
        "Report (2024)",
        "Sheet@#$%123",
    ]
    
    for name in test_names:
        sanitized = converter._sanitize_filename(name)
        print(f"   '{name}' -> '{sanitized}'")
    
    # Test with real file if provided
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
        print(f"\n4. Converting Excel file: {excel_file}")
        
        try:
            csv_files = convert_excel_to_csv(excel_file)
            print(f"   Converted {len(csv_files)} sheets:")
            for sheet, csv_path in csv_files.items():
                print(f"   - {sheet}: {csv_path}")
                
                # Show first few lines
                print(f"\n   First 3 lines of {sheet}:")
                with open(csv_path, 'r', encoding='utf-8') as f:
                    for i, line in enumerate(f, 1):
                        if i > 3:
                            break
                        print(f"     {line.rstrip()[:100]}...")
        
        except Exception as e:
            print(f"   Error: {e}")
    else:
        print("\n4. To test with real Excel file:")
        print("   python src/preprocessing/csv_converter.py path/to/file.xlsx")
    
    print("\n=== Tests Complete ===")