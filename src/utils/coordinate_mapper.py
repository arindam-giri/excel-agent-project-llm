"""
Excel Coordinate Mapping Utilities

Handles conversion between different Excel coordinate systems:
- A1 notation (e.g., "A1", "B5", "AA10")
- R1C1 notation (e.g., "R1C1", "R5C2")
- Zero-based indices (row=0, col=0)
- One-based indices (row=1, col=1)
"""

import re
from typing import Tuple, Optional, Union
from openpyxl.utils import get_column_letter, column_index_from_string
from src.utils.logger import logger


class CoordinateMapper:
    """Convert between different Excel coordinate systems"""
    
    @staticmethod
    def a1_to_indices(cell_ref: str, zero_based: bool = True) -> Tuple[int, int]:
        """
        Convert A1 notation to row/column indices
        
        Args:
            cell_ref: Cell reference like "A1", "B5", "AA10"
            zero_based: If True, return 0-based indices; if False, 1-based
            
        Returns:
            Tuple of (row, col) indices
            
        Examples:
            >>> a1_to_indices("A1")
            (0, 0)
            >>> a1_to_indices("B5", zero_based=False)
            (5, 2)
        """
        # Parse cell reference
        match = re.match(r'^([A-Z]+)(\d+)$', cell_ref.upper())
        if not match:
            raise ValueError(f"Invalid cell reference: {cell_ref}")
        
        col_str, row_str = match.groups()
        
        # Convert to indices
        row = int(row_str)
        col = column_index_from_string(col_str)
        
        if zero_based:
            row -= 1
            col -= 1
        
        return (row, col)
    
    @staticmethod
    def indices_to_a1(row: int, col: int, zero_based: bool = True) -> str:
        """
        Convert row/column indices to A1 notation
        
        Args:
            row: Row index
            col: Column index
            zero_based: If True, inputs are 0-based; if False, 1-based
            
        Returns:
            Cell reference in A1 notation
            
        Examples:
            >>> indices_to_a1(0, 0)
            'A1'
            >>> indices_to_a1(4, 1, zero_based=False)
            'B4'
        """
        if zero_based:
            row += 1
            col += 1
        
        col_letter = get_column_letter(col)
        return f"{col_letter}{row}"
    
    @staticmethod
    def r1c1_to_indices(cell_ref: str, zero_based: bool = True) -> Tuple[int, int]:
        """
        Convert R1C1 notation to row/column indices
        
        Args:
            cell_ref: Cell reference like "R1C1", "R5C2"
            zero_based: If True, return 0-based indices
            
        Returns:
            Tuple of (row, col) indices
            
        Examples:
            >>> r1c1_to_indices("R1C1")
            (0, 0)
            >>> r1c1_to_indices("R5C2", zero_based=False)
            (5, 2)
        """
        match = re.match(r'^R(\d+)C(\d+)$', cell_ref.upper())
        if not match:
            raise ValueError(f"Invalid R1C1 reference: {cell_ref}")
        
        row = int(match.group(1))
        col = int(match.group(2))
        
        if zero_based:
            row -= 1
            col -= 1
        
        return (row, col)
    
    @staticmethod
    def indices_to_r1c1(row: int, col: int, zero_based: bool = True) -> str:
        """
        Convert row/column indices to R1C1 notation
        
        Args:
            row: Row index
            col: Column index
            zero_based: If True, inputs are 0-based
            
        Returns:
            Cell reference in R1C1 notation
            
        Examples:
            >>> indices_to_r1c1(0, 0)
            'R1C1'
            >>> indices_to_r1c1(4, 1, zero_based=False)
            'R4C2'
        """
        if zero_based:
            row += 1
            col += 1
        
        return f"R{row}C{col}"
    
    @staticmethod
    def a1_to_r1c1(cell_ref: str) -> str:
        """
        Convert A1 notation to R1C1 notation
        
        Args:
            cell_ref: Cell reference in A1 notation
            
        Returns:
            Cell reference in R1C1 notation
            
        Examples:
            >>> a1_to_r1c1("A1")
            'R1C1'
            >>> a1_to_r1c1("B5")
            'R5C2'
        """
        row, col = CoordinateMapper.a1_to_indices(cell_ref, zero_based=True)
        return CoordinateMapper.indices_to_r1c1(row, col, zero_based=True)
    
    @staticmethod
    def r1c1_to_a1(cell_ref: str) -> str:
        """
        Convert R1C1 notation to A1 notation
        
        Args:
            cell_ref: Cell reference in R1C1 notation
            
        Returns:
            Cell reference in A1 notation
            
        Examples:
            >>> r1c1_to_a1("R1C1")
            'A1'
            >>> r1c1_to_a1("R5C2")
            'B5'
        """
        row, col = CoordinateMapper.r1c1_to_indices(cell_ref, zero_based=True)
        return CoordinateMapper.indices_to_a1(row, col, zero_based=True)
    
    @staticmethod
    def parse_range(range_str: str) -> Tuple[Tuple[int, int], Tuple[int, int]]:
        """
        Parse Excel range notation to start and end indices
        
        Args:
            range_str: Range like "A1:B5", "R1C1:R5C2"
            
        Returns:
            Tuple of ((start_row, start_col), (end_row, end_col)) - zero-based
            
        Examples:
            >>> parse_range("A1:B5")
            ((0, 0), (4, 1))
            >>> parse_range("R1C1:R5C2")
            ((0, 0), (4, 1))
        """
        if ':' not in range_str:
            raise ValueError(f"Invalid range format: {range_str}")
        
        start_str, end_str = range_str.split(':')
        
        # Detect notation type
        if start_str.upper().startswith('R') and 'C' in start_str.upper():
            # R1C1 notation
            start = CoordinateMapper.r1c1_to_indices(start_str, zero_based=True)
            end = CoordinateMapper.r1c1_to_indices(end_str, zero_based=True)
        else:
            # A1 notation
            start = CoordinateMapper.a1_to_indices(start_str, zero_based=True)
            end = CoordinateMapper.a1_to_indices(end_str, zero_based=True)
        
        return (start, end)
    
    @staticmethod
    def format_range(start_row: int, start_col: int, end_row: int, end_col: int,
                    notation: str = "A1", zero_based: bool = True) -> str:
        """
        Format range from indices
        
        Args:
            start_row: Starting row index
            start_col: Starting column index
            end_row: Ending row index
            end_col: Ending column index
            notation: "A1" or "R1C1"
            zero_based: If True, inputs are 0-based
            
        Returns:
            Range string like "A1:B5" or "R1C1:R5C2"
            
        Examples:
            >>> format_range(0, 0, 4, 1, notation="A1")
            'A1:B5'
            >>> format_range(0, 0, 4, 1, notation="R1C1")
            'R1C1:R5C2'
        """
        if notation.upper() == "A1":
            start = CoordinateMapper.indices_to_a1(start_row, start_col, zero_based)
            end = CoordinateMapper.indices_to_a1(end_row, end_col, zero_based)
        elif notation.upper() == "R1C1":
            start = CoordinateMapper.indices_to_r1c1(start_row, start_col, zero_based)
            end = CoordinateMapper.indices_to_r1c1(end_row, end_col, zero_based)
        else:
            raise ValueError(f"Unknown notation: {notation}")
        
        return f"{start}:{end}"
    
    @staticmethod
    def get_range_dimensions(range_str: str) -> Tuple[int, int]:
        """
        Get dimensions (rows, cols) of a range
        
        Args:
            range_str: Range like "A1:B5"
            
        Returns:
            Tuple of (num_rows, num_cols)
            
        Examples:
            >>> get_range_dimensions("A1:B5")
            (5, 2)
        """
        (start_row, start_col), (end_row, end_col) = CoordinateMapper.parse_range(range_str)
        
        num_rows = end_row - start_row + 1
        num_cols = end_col - start_col + 1
        
        return (num_rows, num_cols)
    
    @staticmethod
    def is_valid_cell_ref(cell_ref: str) -> bool:
        """
        Check if cell reference is valid
        
        Args:
            cell_ref: Cell reference to validate
            
        Returns:
            True if valid
            
        Examples:
            >>> is_valid_cell_ref("A1")
            True
            >>> is_valid_cell_ref("R1C1")
            True
            >>> is_valid_cell_ref("XYZ")
            False
        """
        # Check A1 notation
        if re.match(r'^[A-Z]+\d+$', cell_ref.upper()):
            return True
        
        # Check R1C1 notation
        if re.match(r'^R\d+C\d+$', cell_ref.upper()):
            return True
        
        return False
    
    @staticmethod
    def expand_range(range_str: str) -> list[str]:
        """
        Expand range to list of individual cell references
        
        Args:
            range_str: Range like "A1:B2"
            
        Returns:
            List of cell references
            
        Examples:
            >>> expand_range("A1:B2")
            ['A1', 'B1', 'A2', 'B2']
        """
        (start_row, start_col), (end_row, end_col) = CoordinateMapper.parse_range(range_str)
        
        cells = []
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cells.append(CoordinateMapper.indices_to_a1(row, col, zero_based=True))
        
        return cells
    
    @staticmethod
    def offset_cell(cell_ref: str, row_offset: int, col_offset: int) -> str:
        """
        Offset a cell reference
        
        Args:
            cell_ref: Original cell reference
            row_offset: Number of rows to offset (can be negative)
            col_offset: Number of columns to offset (can be negative)
            
        Returns:
            New cell reference
            
        Examples:
            >>> offset_cell("A1", 1, 1)
            'B2'
            >>> offset_cell("B5", -1, 0)
            'B4'
        """
        row, col = CoordinateMapper.a1_to_indices(cell_ref, zero_based=True)
        new_row = row + row_offset
        new_col = col + col_offset
        
        if new_row < 0 or new_col < 0:
            raise ValueError(f"Offset results in negative indices: {new_row}, {new_col}")
        
        return CoordinateMapper.indices_to_a1(new_row, new_col, zero_based=True)
    
    @staticmethod
    def column_letter_to_index(col_letter: str, zero_based: bool = True) -> int:
        """
        Convert column letter to index
        
        Args:
            col_letter: Column letter like "A", "Z", "AA"
            zero_based: If True, return 0-based index
            
        Returns:
            Column index
            
        Examples:
            >>> column_letter_to_index("A")
            0
            >>> column_letter_to_index("Z", zero_based=False)
            26
        """
        col_idx = column_index_from_string(col_letter)
        
        if zero_based:
            col_idx -= 1
        
        return col_idx
    
    @staticmethod
    def column_index_to_letter(col_index: int, zero_based: bool = True) -> str:
        """
        Convert column index to letter
        
        Args:
            col_index: Column index
            zero_based: If True, input is 0-based
            
        Returns:
            Column letter
            
        Examples:
            >>> column_index_to_letter(0)
            'A'
            >>> column_index_to_letter(26, zero_based=False)
            'Z'
        """
        if zero_based:
            col_index += 1
        
        return get_column_letter(col_index)


class RangeHelper:
    """Helper class for working with Excel ranges"""
    
    @staticmethod
    def merge_ranges(range1: str, range2: str) -> str:
        """
        Merge two ranges into a single range that encompasses both
        
        Args:
            range1: First range
            range2: Second range
            
        Returns:
            Merged range
            
        Examples:
            >>> merge_ranges("A1:B5", "C3:D10")
            'A1:D10'
        """
        (start1_row, start1_col), (end1_row, end1_col) = CoordinateMapper.parse_range(range1)
        (start2_row, start2_col), (end2_row, end2_col) = CoordinateMapper.parse_range(range2)
        
        merged_start_row = min(start1_row, start2_row)
        merged_start_col = min(start1_col, start2_col)
        merged_end_row = max(end1_row, end2_row)
        merged_end_col = max(end1_col, end2_col)
        
        return CoordinateMapper.format_range(
            merged_start_row, merged_start_col,
            merged_end_row, merged_end_col,
            notation="A1", zero_based=True
        )
    
    @staticmethod
    def is_cell_in_range(cell_ref: str, range_str: str) -> bool:
        """
        Check if cell is within a range
        
        Args:
            cell_ref: Cell reference
            range_str: Range to check
            
        Returns:
            True if cell is in range
            
        Examples:
            >>> is_cell_in_range("B3", "A1:C5")
            True
            >>> is_cell_in_range("D6", "A1:C5")
            False
        """
        cell_row, cell_col = CoordinateMapper.a1_to_indices(cell_ref, zero_based=True)
        (start_row, start_col), (end_row, end_col) = CoordinateMapper.parse_range(range_str)
        
        return (start_row <= cell_row <= end_row and
                start_col <= cell_col <= end_col)
    
    @staticmethod
    def split_range_into_chunks(range_str: str, chunk_size: int = 1000) -> list[str]:
        """
        Split large range into smaller chunks (for processing large datasets)
        
        Args:
            range_str: Range to split
            chunk_size: Number of rows per chunk
            
        Returns:
            List of chunk ranges
            
        Examples:
            >>> split_range_into_chunks("A1:B5000", chunk_size=1000)
            ['A1:B1000', 'A1001:B2000', 'A2001:B3000', 'A3001:B4000', 'A4001:B5000']
        """
        (start_row, start_col), (end_row, end_col) = CoordinateMapper.parse_range(range_str)
        
        chunks = []
        current_row = start_row
        
        while current_row <= end_row:
            chunk_end_row = min(current_row + chunk_size - 1, end_row)
            chunk_range = CoordinateMapper.format_range(
                current_row, start_col,
                chunk_end_row, end_col,
                notation="A1", zero_based=True
            )
            chunks.append(chunk_range)
            current_row = chunk_end_row + 1
        
        return chunks


# Convenience functions
def a1_to_row_col(cell_ref: str) -> Tuple[int, int]:
    """Shorthand for a1_to_indices with zero_based=True"""
    return CoordinateMapper.a1_to_indices(cell_ref, zero_based=True)


def row_col_to_a1(row: int, col: int) -> str:
    """Shorthand for indices_to_a1 with zero_based=True"""
    return CoordinateMapper.indices_to_a1(row, col, zero_based=True)


# Export main classes and functions
__all__ = [
    'CoordinateMapper',
    'RangeHelper',
    'a1_to_row_col',
    'row_col_to_a1',
]


if __name__ == "__main__":
    # Test coordinate conversions
    print("=== Coordinate Mapper Tests ===\n")
    
    # A1 to indices
    print("A1 to indices:", CoordinateMapper.a1_to_indices("A1"))
    print("B5 to indices:", CoordinateMapper.a1_to_indices("B5"))
    print("AA10 to indices:", CoordinateMapper.a1_to_indices("AA10"))
    
    # Indices to A1
    print("\nIndices to A1:", CoordinateMapper.indices_to_a1(0, 0))
    print("Indices to A1:", CoordinateMapper.indices_to_a1(4, 1))
    
    # R1C1 conversions
    print("\nA1 to R1C1:", CoordinateMapper.a1_to_r1c1("B5"))
    print("R1C1 to A1:", CoordinateMapper.r1c1_to_a1("R5C2"))
    
    # Range parsing
    print("\nParse range A1:B5:", CoordinateMapper.parse_range("A1:B5"))
    print("Range dimensions:", CoordinateMapper.get_range_dimensions("A1:B5"))
    
    # Range expansion
    print("\nExpand A1:B2:", CoordinateMapper.expand_range("A1:B2"))
    
    # Offset
    print("\nOffset A1 by (1,1):", CoordinateMapper.offset_cell("A1", 1, 1))
    
    # Range operations
    print("\nMerge A1:B5 and C3:D10:", RangeHelper.merge_ranges("A1:B5", "C3:D10"))
    print("Is B3 in A1:C5?", RangeHelper.is_cell_in_range("B3", "A1:C5"))
    
    # Chunking
    chunks = RangeHelper.split_range_into_chunks("A1:B100", chunk_size=30)
    print(f"\nSplit A1:B100 into chunks of 30: {chunks}")